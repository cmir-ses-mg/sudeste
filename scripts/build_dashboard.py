#!/usr/bin/env python3
"""
Gera o index.html (Dashboard + De/Para) a partir da planilha de
Monitoramento Fonte 95 (abas Consolidado + Plano Sudeste).

Uso:
    python3 build_dashboard.py CAMINHO_DO_EXCEL.xlsx SAIDA.html

Este script é a única fonte da verdade da lógica de negócio (equivalente
ao que antes vivia em medidas DAX no Power BI). Qualquer ajuste de regra
de negócio deve ser feito aqui.
"""
import sys
import collections
import openpyxl
import datetime

# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────

def trim(s):
    return (s or '').strip()

def norm(s):
    return trim(s).upper()

def cap(s):
    return s[0].upper() + s[1:].lower() if s else s

def fmt(v, pat):
    if pat == "0.00":
        return f'{v:.2f}'.replace('.', ',')
    if pat == "0.0":
        return f'{v:.1f}'.replace('.', ',')
    if pat == "#,##0.00":
        return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    raise ValueError(pat)

def fmtM(v):
    if abs(v) >= 1e6:
        return 'R$ ' + fmt(v / 1e6, "0.00") + ' mi'
    if abs(v) >= 1e3:
        return 'R$ ' + fmt(v / 1e3, "0.0") + ' mil'
    return 'R$ ' + fmt(v, "#,##0.00")

MICRO_MAP = {
    'Juiz de Fora': 'Juiz de Fora', 'Ubá': 'Ubá', 'Muriaé': 'Muriaé',
    'São João Nepomuceno': 'São João Nepomuceno/Bicas', 'Bicas': 'São João Nepomuceno/Bicas',
    'Leopoldina': 'Leopoldina/Cataguases', 'Cataguases': 'Leopoldina/Cataguases',
    'Santos Dumont': 'Santos Dumont', 'Além Paraíba': 'Além Paraíba', 'Lima Duarte': 'Lima Duarte',
    'Carangola': 'Carangola', 'Toda a macrorregião': 'Transportes em geral',
    'Consórcios generalistas': 'Fortalecimento dos Consórcios',
}

# Município (Consolidado) -> microrregião, e se é o "polo" que casa 1:1
# com um Município Polo do Plano Sudeste (só esses têm "Plano inicial").
MUN_TO_MICRO = {
    'JUIZ DE FORA': ('Juiz de Fora', True), 'MATIAS BARBOSA': ('Juiz de Fora', False),
    'RIO NOVO': ('Juiz de Fora', False), 'UBA': ('Ubá', True), 'MURIAE': ('Muriaé', True),
    'CATAGUASES': ('Leopoldina/Cataguases', True),
}

STATUS_MAP = {
    'Já rodando': 'Em execução', 'Fora': 'Fora do escopo', 'Fora - Desistiram': 'Desistência confirmada',
    'Definir': 'Em definição', 'Rodar': 'Aprovado para execução', 'Aguarda Projeto': 'Aguardando elaboração de projeto',
    'Análise Preliminar': 'Em análise preliminar', 'Melhor em Bicas': 'Sugestão de realocação territorial',
    'Não viável para o momento': 'Inviável no momento',
    'Regional vai pedir ofício para começarmos a rodar.': 'Aguardando ofício regional',
}

INSTR_COR = {'CONVÊNIO': '#1D9E75', 'RESOLUÇÃO': '#378ADD', 'RESOLUÇÃO UBS': '#EF9F27', 'EXECUÇÃO DIRETA': '#888780'}
CARD_BG = {'CONVÊNIO': '#f0faf6', 'RESOLUÇÃO': '#eef5fd', 'RESOLUÇÃO UBS': '#fdf6ea'}
CARD_BD = {'CONVÊNIO': '#9FE1CB', 'RESOLUÇÃO': '#B5D4F4', 'RESOLUÇÃO UBS': '#FAC775'}
CARD_CT = {'CONVÊNIO': '#0F6E56', 'RESOLUÇÃO': '#185FA5', 'RESOLUÇÃO UBS': '#854F0B'}
PILL_CLS = {'PAGO': 'pill-green', 'EM TRAMITAÇÃO': 'pill-blue', 'EM ANÁLISE PRÉVIA': 'pill-amber'}
GRP_ORDER = ['PAGO', 'EM TRAMITAÇÃO', 'EM ANÁLISE PRÉVIA', 'AGUARDA CONVÊNIO PROJETO', 'EM EXECUÇÃO']
GRP_BG = {'PAGO': '#f0faf6', 'EM TRAMITAÇÃO': '#eef5fd', 'EM ANÁLISE PRÉVIA': '#fdf6ea'}
GRP_DOT = {'PAGO': '#1D9E75', 'EM TRAMITAÇÃO': '#378ADD', 'EM ANÁLISE PRÉVIA': '#EF9F27', 'AGUARDA CONVÊNIO PROJETO': '#B4B2A9'}
GRP_TX = {'PAGO': '#0F6E56', 'EM TRAMITAÇÃO': '#185FA5', 'EM ANÁLISE PRÉVIA': '#854F0B'}
ETAPA_COR = {
    'Resolução em cadastramento no SESResolve': '#378ADD',
    'ETAPA 7 - PUBLICADA - EM TRÂMITE DE CADASTRO E ADESÃO NO SES RESOLVE': '#1D9E75',
    'ETAPA 8 - EM TRÂMITE DE PAGAMENTO': '#EF9F27',
    'ETAPA 2/3 - ANÁLISE ASSISTENCIAL/ENGENHARIA - OBRA': '#BA7517',
    'ETAPA 2 - ANÁLISE DA ENGENHARIA - OBRA': '#BA7517',
    'ETAPA 2 - AGUARDANDO ENVIO DE DOCUMENTAÇÃO DO CHECKLIST PARA INICIAR TRAMITAÇÃO': '#B4B2A9',
    'ETAPA 4 - EM ANÁLISE DA DIFE - MUNICÍPIO': '#888780',
    'ETAPA 1 - CADASTRAMENTO DA PROPOSTA': '#5F5E5A',
    'Aguarda elaboração de resolução': '#D85A30',
    'Aguarda emissão de parecer técnico referente à indicação parlamentar': '#A32D2D',
}
LIMIAR_PCT = 10


def san(v):
    s = '' if v is None else str(v).strip()
    return s.replace('\n', ' ').replace('\r', ' ').replace("'", "’")


# ─────────────────────────────────────────────────────────────────────
# Carga dos dados
# ─────────────────────────────────────────────────────────────────────

# Ordem canônica que o resto do script espera de cada linha do Consolidado
# (posição -> nome exato do cabeçalho na planilha). A posição 8 (status) usa
# "prefixo" porque o nome dessa coluna já mudou antes (ex.: "STATUS (03/07)"),
# então localizamos por começar com "STATUS" em vez de exigir o texto exato.
CAMPOS_CONSOLIDADO = [
    ('INSTRUMENTO', 'exato'),
    ('NÚMERO DA INDICAÇÃO', 'exato'),
    ('OBJETO', 'exato'),
    ('MUNICÍPIO', 'exato'),
    ('BENEFICIÁRIO', 'exato'),
    (None, None),  # posição não utilizada pelo restante do script
    ('VALOR DO PLEITO', 'exato'),
    ('PADRINHO', 'exato'),  # não é exibido, mas mantido na posição por compatibilidade
    ('STATUS', 'prefixo'),
    ('ETAPA DE CELEBRAÇÃO', 'exato'),
    ('DETALHAMENTO ETAPA DE CELEBRAÇÃO', 'exato'),
    ('SEI', 'exato'),
    ('FONTE FINAL', 'exato'),
    ('ACORDO SUDESTE', 'exato'),
]


def remapear_consolidado(ws):
    """Lê a aba Consolidado pelo NOME do cabeçalho (não pela posição da
    coluna), e devolve cada linha já reordenada na ordem canônica que o
    resto do script espera. Assim, se alguém reordenar, inserir ou remover
    uma coluna na planilha, o script continua funcionando — só quebra de
    verdade se uma coluna essencial for renomeada de um jeito irreconhecível."""
    headers = [norm(c.value) for c in ws[1]]

    def localizar(nome, modo):
        if nome is None:
            return None
        if modo == 'exato':
            for i, h in enumerate(headers):
                if h == nome:
                    return i
        elif modo == 'prefixo':
            for i, h in enumerate(headers):
                if h.startswith(nome):
                    return i
        return None

    indices = [localizar(nome, modo) for nome, modo in CAMPOS_CONSOLIDADO]

    essenciais = {'MUNICÍPIO': 3, 'VALOR DO PLEITO': 6, 'STATUS': 8}
    faltando = [nome for nome, pos in essenciais.items() if indices[pos] is None]
    if faltando:
        raise RuntimeError(
            f"Não encontrei a(s) coluna(s) essencial(is) {faltando} na aba Consolidado. "
            f"Cabeçalhos encontrados: {[c.value for c in ws[1]]}"
        )

    linhas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nova = tuple(row[i] if i is not None and i < len(row) else None for i in indices)
        linhas.append(nova)
    return linhas


def carregar_dados(caminho_xlsx):
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)

    cons_all = [r for r in remapear_consolidado(wb['Consolidado']) if r[3]]
    cons = [r for r in cons_all if norm(r[13]) not in ('NÃO', 'NAO')]

    plano = [r for r in wb['Plano Sudeste'].iter_rows(min_row=2, values_only=True) if r[0]]
    plano_micro = []
    for r in plano:
        micro = MICRO_MAP.get(trim(r[0]))
        plano_micro.append({
            'micro': micro, 'dest': trim(r[1]), 'val': r[2] or 0,
            'status': trim(r[4]), 'obs': trim(r[6]) if len(r) > 6 else '',
        })
    return cons, plano_micro


# ─────────────────────────────────────────────────────────────────────
# PÁGINA 1 — Dashboard
# ─────────────────────────────────────────────────────────────────────

def gerar_dashboard(cons, plano_micro, data_ref):
    ST = lambda r: trim(r[8])
    VAL = lambda r: float(r[6] or 0)

    total = sum(VAL(r) for r in cons)
    muns = collections.defaultdict(list)
    for r in cons:
        muns[trim(r[3])].append(r)
    num_ind, num_mun = len(cons), len(muns)
    previsao = sum(r['val'] for r in plano_micro)

    vp = sum(VAL(r) for r in cons if ST(r) == 'PAGO')
    va = sum(VAL(r) for r in cons if ST(r) == 'EM ANÁLISE PRÉVIA')
    vg = sum(VAL(r) for r in cons if ST(r) == 'AGUARDA CONVÊNIO PROJETO')
    vt_literal = sum(VAL(r) for r in cons if ST(r) == 'EM TRAMITAÇÃO')
    cp = sum(1 for r in cons if ST(r) == 'PAGO')
    ct_literal = sum(1 for r in cons if ST(r) == 'EM TRAMITAÇÃO')
    can_al = sum(1 for r in cons if ST(r) in ('EM ANÁLISE PRÉVIA', 'AGUARDA CONVÊNIO PROJETO'))
    vanaltot = va + vg

    p_pago_prev = vp / previsao * 100 if previsao else 0
    p_tram_prev = vanaltot / previsao * 100 if previsao else 0
    p_pleito_prev = total / previsao * 100 if previsao else 0

    # ── Por instrumento ──
    instr_html = ''
    for i, cor in INSTR_COR.items():
        t_ = [r for r in cons if trim(r[0]) == i]
        if not t_:
            continue
        v = sum(VAL(r) for r in t_)
        p = v / total * 100 if total else 0
        instr_html += (
            f"<div class='instr-row'><div class='instr-dot' style='background:{cor}'></div>"
            f"<div class='instr-nome'>{cap(i)}</div><div class='instr-cnt'>{len(t_)}</div>"
            f"<div class='instr-bar-wrap'><div class='instr-bar-fill' style='width:{fmt(p,'0.0')}%;background:{cor}'></div></div>"
            f"<div class='instr-val'>{fmtM(v)}</div><div class='instr-pct'>{fmt(p,'0.0')}%</div></div>"
        )

    # ── Por status ──
    st_grid = [('PAGO', '#1D9E75'), ('EM TRAMITAÇÃO', '#378ADD'), ('EM ANÁLISE PRÉVIA', '#EF9F27'),
               ('EM EXECUÇÃO', '#888780'), ('AGUARDA CONVÊNIO PROJETO', '#B4B2A9')]
    grid_html = ''
    for s, cor in st_grid:
        t_ = [r for r in cons if ST(r) == s]
        if not t_:
            continue
        v = sum(VAL(r) for r in t_)
        grid_html += (
            f"<div class='st-pill'><div class='st-dot' style='background:{cor}'></div>"
            f"<div class='st-info'><div class='st-nome'>{cap(s)}</div><div class='st-val'>{fmtM(v)}</div>"
            f"<div class='st-cnt'>{len(t_)} indicações</div></div></div>"
        )

    # ── Municípios — NOVA estrutura em 4 colunas ──────────────────────
    # Colunas: Plano inicial | Pleito mapeado / Validado | Em tramitação | Pago/Execução
    mun_html = ''
    for mun, items in sorted(muns.items(), key=lambda kv: -sum(VAL(r) for r in kv[1])):
        mval = sum(VAL(r) for r in items)
        sp = mun.find(' ')
        ini = (mun[0] + (mun[sp + 1] if 0 < sp < len(mun) - 1 else '')).upper()

        micro_info = MUN_TO_MICRO.get(mun.upper())
        # soma da previsão da microrregião correspondente (só no polo)
        plano_inicial_txt = '—'
        if micro_info and micro_info[1]:
            micro_nome = micro_info[0]
            prev_micro = sum(r['val'] for r in plano_micro if r['micro'] == micro_nome)
            plano_inicial_txt = fmtM(prev_micro) if prev_micro else '—'

        v_pago = sum(VAL(r) for r in items if ST(r) == 'PAGO')
        v_tram = sum(VAL(r) for r in items if ST(r) in ('EM TRAMITAÇÃO', 'EM ANÁLISE PRÉVIA', 'AGUARDA CONVÊNIO PROJETO', 'EM EXECUÇÃO'))
        v_mapeado = mval  # pleito mapeado / validado = tudo que está no Consolidado p/ esse município

        mini = ''
        seen = []
        for r in items:
            ins = trim(r[0])
            if ins not in seen:
                seen.append(ins)
        for ins in seen:
            iv = sum(VAL(r) for r in items if trim(r[0]) == ins)
            mini += f"<div class='bmini-seg' style='width:{fmt(iv/mval*100,'0.0')}%;background:{INSTR_COR.get(ins,'#888')}'></div>"

        # ── detalhe expansível: mesma lógica de antes, agrupado por status literal ──
        present = [s for s in GRP_ORDER if any(ST(r) == s for r in items)]
        grupos = ''
        for idx_s, s in enumerate(present):
            g = sorted([r for r in items if ST(r) == s], key=lambda r: -VAL(r))
            gv = sum(VAL(r) for r in g)
            open_a = ' open' if idx_s == 0 else ''
            cards = ''
            for r in g:
                instr = san(r[0]); numero = san(r[1]); numero = '' if numero.upper() == 'NA' else numero
                obj, ben, stt = san(r[2]), san(r[4]), san(r[8])
                etp, det, sei, fon, sud = san(r[9]), san(r[10]), san(r[11]), san(r[12]), norm(r[13])
                etapa_txt = etp if etp and etp != 'PAGO' and etp != stt else ''
                det_txt = det if det and det != stt and det != 'PAGO' else ''
                cbg = CARD_BG.get(instr, '#f5f5f4'); cbd = CARD_BD.get(instr, '#D3D1C7'); cct = CARD_CT.get(instr, '#5F5E5A')
                stp = PILL_CLS.get(stt, 'pill-gray')
                c = (f"<div class='acao-card' style='background:{cbg};border-color:{cbd}'>"
                     f"<div class='acao-top'><span class='acao-tipo' style='color:{cct}'>{instr}</span>"
                     f"<span class='acao-valor' style='color:{cct}'>R$ {fmt(VAL(r),'#,##0.00')}</span></div>"
                     f"<div class='acao-titulo'>{obj}</div><div class='acao-sep' style='background:{cbd}'></div>"
                     f"<div class='acao-row'>"
                     f"<div class='acao-item'><span class='acao-lbl'>Beneficiário</span><span class='acao-v' title='{ben}'>{ben}</span></div>"
                     f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Status</span><span class='pill {stp}' style='font-size:10px;padding:2px 8px'>{stt}</span></div>")
                if numero: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Nº indicação</span><span class='acao-v'>{numero}</span></div>"
                if etapa_txt: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Etapa</span><span class='acao-v' title='{etapa_txt}'>{etapa_txt}</span></div>"
                if det_txt: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Detalhe</span><span class='acao-v' title='{det_txt}'>{det_txt}</span></div>"
                if sei: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>SEI</span><span class='acao-v'>{sei}</span></div>"
                if fon: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Fonte</span><span class='acao-v'>{fon}</span></div>"
                if sud == 'SIM': c += "<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Acordo Sudeste</span><span class='pill pill-blue' style='font-size:10px;padding:2px 8px'>SIM</span></div>"
                cards += c + '</div></div>'
            gtx = GRP_TX.get(s, '#5F5E5A')
            grupos += (
                f"<details class='st-grupo' style='border-color:{GRP_DOT.get(s,'#888780')}33'{open_a}>"
                f"<summary class='st-grupo-hdr' style='background:{GRP_BG.get(s,'#f5f5f4')}'>"
                f"<div class='st-grupo-left'><div class='st-grupo-dot' style='background:{GRP_DOT.get(s,'#888780')}'></div>"
                f"<span class='st-grupo-nome' style='color:{gtx}'>{cap(s)}</span></div>"
                f"<div class='st-grupo-right'><span class='st-grupo-val' style='color:{gtx}'>{fmtM(gv)}</span>"
                f"<span class='st-grupo-cnt' style='color:{gtx}'>{len(g)} {'indicação' if len(g)==1 else 'indicações'}</span>"
                f"<span class='st-grupo-chv'>›</span></div></summary><div class='st-grupo-body'>{cards}</div></details>"
            )

        mun_html += (
            f"<details class='brow'><summary class='bhead mun-cols'>"
            f"<div class='bleft'><div class='bav'>{ini}</div><div class='binfo'><div class='bnome'>{cap(mun)}</div>"
            f"<div class='bmini-bar'>{mini}</div></div></div>"
            f"<div class='mun-col'><div class='mun-col-lbl'>Plano inicial</div><div class='mun-col-val'>{plano_inicial_txt}</div></div>"
            f"<div class='mun-col'><div class='mun-col-lbl'>Pleito mapeado</div><div class='mun-col-val' style='color:#185FA5'>{fmtM(v_mapeado)}</div></div>"
            f"<div class='mun-col'><div class='mun-col-lbl'>Em tramitação</div><div class='mun-col-val' style='color:#8A5A17'>{fmtM(v_tram) if v_tram else '—'}</div></div>"
            f"<div class='mun-col'><div class='mun-col-lbl'>Pago/Execução</div><div class='mun-col-val' style='color:#0F6E56'>{fmtM(v_pago) if v_pago else '—'}</div></div>"
            f"<div class='bchevron'>›</div></summary><div class='bdet'><div class='acao-lista'>{grupos}</div></div></details>"
        )

    # ── Etapas de celebração — REMOVIDO por pedido do usuário ──
    # (mantido no script por eventual necessidade futura, mas não gerado no HTML)
    etapas = collections.defaultdict(list)
    for r in cons:
        if ST(r) != 'EM TRAMITAÇÃO':
            continue
        etapas[trim(r[9])].append(r)

    html = f"""<div class='top'>
  <div class='fonte-badge'>Fonte 95 — Emendas Parlamentares</div>
  <div class='top-titulo'>Monitoramento Consolidado</div>
  <div class='top-sub'>Referência: {data_ref} · {num_ind} indicações · {num_mun} municípios</div>
  <div class='tot-lbl'>Previsão inicial (Plano Sudeste)</div>
  <div class='tot-val'>{fmtM(previsao)}</div>
  <div class='exec-hdr' style='margin-top:16px'>
    <span class='exec-titulo'>Distribuição por status (% da previsão)</span>
    <span class='exec-pct'>{fmt(p_pago_prev,'0.0')}% pago</span>
  </div>
  <div class='exec-bar-wrap'>
    <div style='position:absolute;top:0;left:0;width:{fmt(p_tram_prev,'0.00').replace(',','.')}%;height:100%;background:#EF9F27;z-index:1;'></div>
    <div style='position:absolute;top:0;left:0;width:{fmt(p_pago_prev,'0.00').replace(',','.')}%;height:100%;background:#1D9E75;z-index:2;'></div>
    <div style='position:absolute;top:0;left:0;width:{fmt(p_pleito_prev,'0.00').replace(',','.')}%;height:100%;border-top:2.5px dashed #378ADD;border-bottom:2.5px dashed #378ADD;z-index:3;'></div>
  </div>
  <div class='exec-legenda'>
    <div class='leg-item'><div class='leg-dot' style='background:#f1f5f9;border:0.5px solid #e2e8f0'></div>Previsão inicial</div>
    <div class='leg-item'><div class='leg-dot' style='background:#378ADD'></div>Pleitos mapeados</div>
    <div class='leg-item'><div class='leg-dot' style='background:#EF9F27'></div>Em tramitação</div>
    <div class='leg-item'><div class='leg-dot' style='background:#1D9E75'></div>Pago</div>
  </div>
</div>

<div class='kpi-grid'>
  <div class='kpi'>
    <div class='kpi-lbl'>Previsão inicial</div>
    <div class='kpi-val' style='color:#5F5E5A;'>{fmtM(previsao)}</div>
    <div class='kpi-sub'>Plano Sudeste</div>
  </div>
  <div class='kpi'>
    <div class='kpi-lbl'>Pleitos mapeados</div>
    <div class='kpi-val' style='color:#378ADD;'>{fmtM(total)}</div>
    <div class='kpi-sub'>{num_ind} indicações</div>
  </div>
  <div class='kpi'>
    <div class='kpi-lbl'>Em tramitação</div>
    <div class='kpi-val' style='color:#BA7517;'>{fmtM(vanaltot)}</div>
    <div class='kpi-sub'>{can_al} indicações</div>
  </div>
  <div class='kpi'>
    <div class='kpi-lbl'>Pago</div>
    <div class='kpi-val' style='color:#1D9E75;'>{fmtM(vp)}</div>
    <div class='kpi-sub'>{cp} indicações</div>
  </div>
</div>

<div class='two-col'>
  <div class='card' style='margin-bottom:0'>
    <div class='card-hdr'>
      <span class='card-titulo'>Por instrumento</span>
      <span class='pill pill-blue'>{num_ind} indicações</span>
    </div>
    <div>{instr_html}</div>
  </div>
  <div class='card' style='margin-bottom:0'>
    <div class='card-hdr'>
      <span class='card-titulo'>Por status</span>
    </div>
    <div class='st-grid'>{grid_html}</div>
  </div>
</div>

<div class='card' style='margin-top:14px'>
  <div class='card-hdr'>
    <span class='card-titulo'>Municípios e indicações</span>
    <span class='pill pill-blue'>{num_mun} municípios</span>
  </div>
  <div class='mun-cols-head'>
    <div></div>
    <div>Plano inicial</div>
    <div>Pleito mapeado</div>
    <div>Em tramitação</div>
    <div>Pago/Execução</div>
    <div></div>
  </div>
  <div>{mun_html}</div>
</div>
"""
    return html


# ─────────────────────────────────────────────────────────────────────
# PÁGINA 2 — De/Para
# ─────────────────────────────────────────────────────────────────────

def gerar_depara(cons, plano_micro, data_ref):
    dot = collections.defaultdict(float)
    for r in plano_micro:
        dot[r['micro']] += r['val']

    def grp(muns):
        return [r for r in cons if trim(r[3]) in muns]

    def s(t, pago=False):
        return sum(float(r[6] or 0) for r in t if not pago or trim(r[8]) == 'PAGO')

    JF, UB, MU, LE = grp({'JUIZ DE FORA', 'MATIAS BARBOSA', 'RIO NOVO'}), grp({'UBA'}), grp({'MURIAE'}), grp({'CATAGUASES'})
    ple = {'Juiz de Fora': s(JF), 'Ubá': s(UB), 'Muriaé': s(MU), 'Leopoldina/Cataguases': s(LE)}
    pag = {'Juiz de Fora': s(JF, 1), 'Ubá': s(UB, 1), 'Muriaé': s(MU, 1), 'Leopoldina/Cataguases': s(LE, 1)}
    exec_tbl = {'Juiz de Fora': JF, 'Ubá': UB, 'Muriaé': MU, 'Leopoldina/Cataguases': LE}

    de_total = sum(dot.values())
    para_total = sum(ple.values())
    pago_total = sum(pag.values())
    diff_total = para_total - de_total
    diff_total_txt = ('+' if diff_total >= 0 else '−') + 'R$ ' + fmt(abs(diff_total) / 1e6, "0.00") + ' mi'

    # cores dos grupos de status, no MESMO padrão visual do Dashboard
    # ("Previsão" é um grupo extra, cor neutra/roxa para diferenciar de execução)
    grp_bg_dp = {'PREVISAO': '#F5F3FF', **GRP_BG}
    grp_dot_dp = {'PREVISAO': '#8B5CF6', **GRP_DOT}
    grp_tx_dp = {'PREVISAO': '#5B21B6', **GRP_TX}

    rows = ''
    for m in sorted(dot, key=lambda x: -dot[x]):
        de_v = dot[m]; para_v = ple.get(m, 0); pago_v = pag.get(m, 0)
        diff = para_v - de_v
        pct = diff / de_v * 100 if de_v else 0
        de_txt = fmtM(de_v)
        para_txt = '—' if para_v == 0 else fmtM(para_v)
        pago_txt = '—' if pago_v == 0 else fmtM(pago_v)
        if para_v == 0:
            diff_txt = '—'
        else:
            diff_txt = ('▲ +' if diff >= 0 else '▼ −') + 'R$ ' + fmt(abs(diff) / 1e6, "0.00") + ' mi (' + ('+' if diff >= 0 else '−') + fmt(abs(pct), "0.0") + '%)'
        if para_v == 0:
            sit, sit_bg, sit_cor, diff_cor = 'Ainda não mapeado', '#F1EFE8', '#5F5E5A', '#94a3b8'
        elif abs(pct) <= LIMIAR_PCT:
            sit, sit_bg, sit_cor, diff_cor = 'Dentro do previsto', '#E1F5EE', '#0F6E56', '#64748b'
        elif pct > LIMIAR_PCT:
            sit, sit_bg, sit_cor, diff_cor = 'Acima do previsto', '#FDF0DD', '#8A5A17', '#BA7517'
        else:
            sit, sit_bg, sit_cor, diff_cor = 'Abaixo do previsto', '#FCEBEB', '#791F1F', '#C0392B'

        itens_plan = sorted([r for r in plano_micro if r['micro'] == m], key=lambda r: -r['val'])
        itens_exec = exec_tbl.get(m, [])

        # ── painel "Planejado (Plano Sudeste)" — agrupado por status, cada grupo expansível ──
        plan_por_status = collections.defaultdict(list)
        for r in itens_plan:
            st_p = r['status']; dest_upper = norm(r['dest'])
            match_exec = [c for c in itens_exec if dest_upper in norm(c[2]) or norm(c[2]) in dest_upper]
            remapeado = st_p.startswith('Fora') and len(match_exec) > 0
            label_final = 'Remapeado' if remapeado else STATUS_MAP.get(st_p, st_p)
            plan_por_status[label_final].append((r, remapeado))

        # cor de cada grupo dentro do painel Planejado (por rótulo já traduzido)
        plan_grp_cor = {
            'Remapeado': ('#EDE9FE', '#5B21B6'), 'Em execução': ('#E1F5EE', '#0F6E56'),
            'Fora do escopo': ('#FCEBEB', '#791F1F'), 'Desistência confirmada': ('#FCEBEB', '#791F1F'),
            'Em definição': ('#F1EFE8', '#5F5E5A'),
        }
        plan_groups_html = ''
        for idx_lbl, (label, itens_lbl) in enumerate(sorted(plan_por_status.items(), key=lambda kv: -sum(r['val'] for r, _ in kv[1]))):
            bg, tx = plan_grp_cor.get(label, ('#FDF0DD', '#8A5A17'))
            gv = sum(r['val'] for r, _ in itens_lbl)
            cards_lbl = ''
            for r, remapeado in itens_lbl:
                obs_html = f" <span class='dp-item-obs'>{r['obs']}</span>" if r['obs'] else ''
                cards_lbl += (
                    f"<div class='dp-item'><div class='dp-item-top'><span class='dp-item-dest'>{r['dest']}</span>"
                    f"<span class='dp-item-val'>R$ {fmt(r['val'],'#,##0.00')}</span></div>"
                    f"<div class='dp-item-meta'>{obs_html}</div></div>"
                )
            plan_groups_html += (
                f"<details class='st-grupo' style='border-color:{tx}33'{' open' if idx_lbl == 0 else ''}>"
                f"<summary class='st-grupo-hdr' style='background:{bg}'>"
                f"<div class='st-grupo-left'><div class='st-grupo-dot' style='background:{tx}'></div>"
                f"<span class='st-grupo-nome' style='color:{tx}'>{label}</span></div>"
                f"<div class='st-grupo-right'><span class='st-grupo-val' style='color:{tx}'>{fmtM(gv)}</span>"
                f"<span class='st-grupo-cnt' style='color:{tx}'>{len(itens_lbl)} {'item' if len(itens_lbl)==1 else 'itens'}</span>"
                f"<span class='st-grupo-chv'>›</span></div></summary><div class='st-grupo-body'>{cards_lbl}</div></details>"
            )

        # ── painel "Em execução (Consolidado)" — agrupado por status literal, cada grupo expansível ──
        present = [st for st in GRP_ORDER if any(trim(r[8]) == st for r in itens_exec)]
        exec_groups_html = ''
        for idx_s, st in enumerate(present):
            g = sorted([r for r in itens_exec if trim(r[8]) == st], key=lambda r: -float(r[6] or 0))
            gv = sum(float(r[6] or 0) for r in g)
            cards_exec = ''
            for r in g:
                instr = san(r[0]); numero = san(r[1]); numero = '' if numero.upper() == 'NA' else numero
                obj, ben, stt = san(r[2]), san(r[4]), san(r[8])
                etp, det, sei, fon, sud = san(r[9]), san(r[10]), san(r[11]), san(r[12]), norm(r[13])
                etapa_txt = etp if etp and etp != 'PAGO' and etp != stt else ''
                det_txt = det if det and det != stt and det != 'PAGO' else ''
                cbg = CARD_BG.get(instr, '#f5f5f4'); cbd = CARD_BD.get(instr, '#D3D1C7'); cct = CARD_CT.get(instr, '#5F5E5A')
                stp = PILL_CLS.get(stt, 'pill-gray')
                c = (f"<div class='acao-card' style='background:{cbg};border-color:{cbd}'>"
                     f"<div class='acao-top'><span class='acao-tipo' style='color:{cct}'>{instr}</span>"
                     f"<span class='acao-valor' style='color:{cct}'>R$ {fmt(float(r[6] or 0),'#,##0.00')}</span></div>"
                     f"<div class='acao-titulo'>{obj}</div><div class='acao-sep' style='background:{cbd}'></div>"
                     f"<div class='acao-row'>"
                     f"<div class='acao-item'><span class='acao-lbl'>Beneficiário</span><span class='acao-v' title='{ben}'>{ben}</span></div>"
                     f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Status</span><span class='pill {stp}' style='font-size:10px;padding:2px 8px'>{stt}</span></div>")
                if numero: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Nº indicação</span><span class='acao-v'>{numero}</span></div>"
                if etapa_txt: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Etapa</span><span class='acao-v' title='{etapa_txt}'>{etapa_txt}</span></div>"
                if det_txt: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Detalhe</span><span class='acao-v' title='{det_txt}'>{det_txt}</span></div>"
                if sei: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>SEI</span><span class='acao-v'>{sei}</span></div>"
                if fon: c += f"<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Fonte</span><span class='acao-v'>{fon}</span></div>"
                if sud == 'SIM': c += "<div class='acao-div'>|</div><div class='acao-item'><span class='acao-lbl'>Acordo Sudeste</span><span class='pill pill-blue' style='font-size:10px;padding:2px 8px'>SIM</span></div>"
                cards_exec += c + '</div></div>'
            gtx = GRP_TX.get(st, '#5F5E5A')
            exec_groups_html += (
                f"<details class='st-grupo' style='border-color:{GRP_DOT.get(st,'#888780')}33'{' open' if idx_s == 0 else ''}>"
                f"<summary class='st-grupo-hdr' style='background:{GRP_BG.get(st,'#f5f5f4')}'>"
                f"<div class='st-grupo-left'><div class='st-grupo-dot' style='background:{GRP_DOT.get(st,'#888780')}'></div>"
                f"<span class='st-grupo-nome' style='color:{gtx}'>{cap(st)}</span></div>"
                f"<div class='st-grupo-right'><span class='st-grupo-val' style='color:{gtx}'>{fmtM(gv)}</span>"
                f"<span class='st-grupo-cnt' style='color:{gtx}'>{len(g)} {'indicação' if len(g)==1 else 'indicações'}</span>"
                f"<span class='st-grupo-chv'>›</span></div></summary><div class='st-grupo-body'>{cards_exec}</div></details>"
            )
        if not itens_exec:
            exec_groups_html = "<div class='dp-vazio'>Nenhuma indicação mapeada ainda na Fonte 95.</div>"

        cnt_plan, cnt_exec = len(itens_plan), len(itens_exec)
        rows += (
            f"<details class='dp-item-wrap'><summary class='dp-row'>"
            f"<div class='dp-cell dp-nome'><span class='dp-chevron'>›</span>{m}</div>"
            f"<div class='dp-cell dp-num'>{de_txt}</div><div class='dp-cell dp-num'>{para_txt}</div>"
            f"<div class='dp-cell dp-num' style='color:{diff_cor};font-weight:600;'>{diff_txt}</div>"
            f"<div class='dp-cell dp-num'>{pago_txt}</div>"
            f"<div class='dp-cell dp-sit'><span class='pill' style='background:{sit_bg};color:{sit_cor};font-size:10px;padding:3px 9px'>{sit}</span></div>"
            f"</summary>"
            f"<div class='dp-detail'>"
            f"<details class='dp-panel'><summary class='dp-panel-hdr'><span>Planejado (Plano Sudeste) · {cnt_plan} {'item' if cnt_plan==1 else 'itens'}</span><span class='dp-panel-chv'>›</span></summary>"
            f"<div class='dp-panel-body'>{plan_groups_html}</div></details>"
            f"<details class='dp-panel'><summary class='dp-panel-hdr'><span>Em execução (Consolidado) · {cnt_exec} {'indicação' if cnt_exec==1 else 'indicações'}</span><span class='dp-panel-chv'>›</span></summary>"
            f"<div class='dp-panel-body'>{exec_groups_html}</div></details>"
            f"</div>"
            f"</details>"
        )

    html = f"""<div class='top'>
  <div class='top-titulo' style='text-align:center'>Previsão × Execução</div>
  <div class='resumo-grid' style='margin-top:18px;'>
    <div class='resumo-item'>
      <div class='resumo-lbl'>Previsão</div>
      <div class='resumo-val' style='color:#5F5E5A'>{fmtM(de_total)}</div>
    </div>
    <div class='resumo-item'>
      <div class='resumo-lbl'>Execução</div>
      <div class='resumo-val' style='color:#378ADD'>{fmtM(para_total)}</div>
    </div>
    <div class='resumo-item'>
      <div class='resumo-lbl'>Diferença</div>
      <div class='resumo-val' style='color:{"#BA7517" if diff_total >= 0 else "#C0392B"}'>{diff_total_txt}</div>
    </div>
    <div class='resumo-item'>
      <div class='resumo-lbl'>Pago</div>
      <div class='resumo-val' style='color:#1D9E75'>{fmtM(pago_total)}</div>
    </div>
  </div>
  <div class='top-sub' style='text-align:center;margin:12px 0 0;'>Referência: {data_ref} · Previsão: desenho inicial · Execução: Consolidado da execução</div>
</div>

<div class='card'>
  <div class='card-hdr'>
    <span class='card-titulo'>previsto × executado (por microrregião/temática)</span>
    <span class='pill pill-blue'>{len(dot)} linhas do plano</span>
  </div>
  <div class='dp-head'>
    <div>Microrregião / Temática</div>
    <div>Previsão</div>
    <div>Executado</div>
    <div>Diferença</div>
    <div>Pago</div>
    <div>Status</div>
  </div>
  <div>{rows}</div>
  <div class='legenda-sit'>
    <span><span class='leg-dot' style='background:#B4B2A9'></span>Ainda não mapeado no monitoramento</span>
    <span><span class='leg-dot' style='background:#1D9E75'></span>Dentro do previsto (±{LIMIAR_PCT}%)</span>
    <span><span class='leg-dot' style='background:#EF9F27'></span>Acima do previsto</span>
    <span><span class='leg-dot' style='background:#C0392B'></span>Abaixo do previsto</span>
  </div>
</div>

<div class='card' style='background:#f8fafc;'>
  <div class='nota'><b>Notas metodológicas: Previsão</b> é o valor previsto para cada microrregião/temática inicialmente no Plano Sudeste. <b>Executado</b> é o que está sendo executado de fato dentro do Plano Sudeste. <b>"Executado"</b> soma todas as indicações da Fonte 95 já mapeadas para aquela microrregião, e "Pago" soma apenas as indicações com status "Pago". A coluna "Mudou?" compara Executado com Previsão, variações de até {LIMIAR_PCT}% são consideradas dentro do previsto; acima disso, sinalizamos se a execução está rodando acima ou abaixo do que foi originalmente desenhado. Microrregiões sem nenhuma indicação no Consolidado aparecem como "ainda não mapeadas" — o plano existe, mas a Fonte 95 ainda não iniciou o acompanhamento daquele investimento. Juiz de Fora inclui Matias Barbosa e Rio Novo.</div>
</div>
"""
    return html


# ─────────────────────────────────────────────────────────────────────
# Montagem final do HTML (CSS + navegação + fundo do mapa)
# ─────────────────────────────────────────────────────────────────────

CSS = """
  * { box-sizing: border-box; margin: 0; padding: 0; }
  html, body { min-height: 100%; }
  body {
    font-family: Segoe UI, sans-serif; color: #1e293b; font-size: 14px;
    background-color: #ffffff;
    background-image: url('assets/sudeste_map.webp');
    background-repeat: no-repeat;
    background-position: center 90px;
    background-size: min(1400px, 92%) auto;
    background-attachment: fixed;
  }
  body::before {
    content: ''; position: fixed; inset: 0;
    background: rgba(255, 255, 255, 0.12); pointer-events: none; z-index: 0;
  }
  .page { max-width: 1180px; margin: 0 auto; padding: 15px 10px; position: relative; z-index: 1; }
  .top { background: rgba(255,255,255,0.55); backdrop-filter: blur(4px); -webkit-backdrop-filter: blur(4px); border: 0.5px solid #e2e8f0; border-radius: 12px; padding: 10px 14px; margin-bottom: 14px; }
  .fonte-badge { display: inline-block; font-size: 11px; font-weight: 500; padding: 3px 11px; border-radius: 12px; background: #E6F1FB; color: #0C447C; margin-bottom: 8px; }
  .top-titulo { font-size: 20px; font-weight: 500; color: #1e293b; margin-bottom: 2px; }
  .top-sub { font-size: 12px; color: #94a3b8; margin-bottom: 14px; }
  .tot-lbl { font-size: 11px; color: #94a3b8; margin-bottom: 3px; }
  .tot-val { font-size: 26px; font-weight: 500; color: #1e293b; }
  .exec-hdr { display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 6px; }
  .exec-titulo { font-size: 11px; color: #94a3b8; }
  .exec-pct { font-size: 16px; font-weight: 700; color: #1D9E75; }
  .exec-bar-wrap { height: 20px; background: #f1f5f9; border-radius: 6px; overflow: hidden; position: relative; margin-bottom: 6px; }
  .exec-legenda { display: flex; flex-wrap: wrap; gap: 12px; font-size: 10px; }
  .leg-item { display: flex; align-items: center; gap: 5px; color: #64748b; }
  .leg-dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }
  .pill { font-size: 11px; font-weight: 500; padding: 2px 8px; border-radius: 12px; white-space: nowrap; }
  .pill-blue { background: #E6F1FB; color: #0C447C; }
  .pill-green { background: #E1F5EE; color: #0F6E56; }
  .pill-amber { background: #FAEEDA; color: #633806; }
  .pill-gray { background: #F1EFE8; color: #5F5E5A; }
  .kpi-grid { display: grid; grid-template-columns: repeat(4,1fr); gap: 10px; margin-bottom: 14px; }
  .kpi { background: rgba(255,255,255,0.55); backdrop-filter: blur(4px); -webkit-backdrop-filter: blur(4px); border: 0.5px solid #e2e8f0; border-radius: 10px; padding: 13px 16px; }
  .kpi-lbl { font-size: 10px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 5px; }
  .kpi-val { font-size: 19px; font-weight: 500; }
  .kpi-sub { font-size: 10px; color: #94a3b8; margin-top: 3px; }
  .card { background: rgba(255,255,255,0.55); backdrop-filter: blur(4px); -webkit-backdrop-filter: blur(4px); border: 0.5px solid #e2e8f0; border-radius: 12px; padding: 16px 20px; margin-bottom: 14px; }
  .card-hdr { display: flex; justify-content: space-between; align-items: center; margin-bottom: 4px; padding-bottom: 10px; border-bottom: 0.5px solid #f1f5f9; }
  .card-titulo { font-size: 11px; font-weight: 500; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.07em; }
  .two-col { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 14px; }
  .instr-row { display: flex; align-items: center; gap: 10px; padding: 9px 0; border-bottom: 0.5px solid #f1f5f9; }
  .instr-row:last-child { border-bottom: none; }
  .instr-dot { width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0; }
  .instr-nome { flex: 1; font-size: 12px; color: #1e293b; }
  .instr-cnt { font-size: 11px; padding: 2px 8px; border-radius: 10px; background: #f8fafc; color: #64748b; border: 0.5px solid #e2e8f0; }
  .instr-val { font-size: 12px; font-weight: 500; color: #1e293b; text-align: right; min-width: 110px; }
  .instr-pct { font-size: 10px; color: #94a3b8; min-width: 38px; text-align: right; }
  .instr-bar-wrap { width: 70px; height: 5px; background: #f1f5f9; border-radius: 3px; overflow: hidden; }
  .instr-bar-fill { height: 100%; border-radius: 3px; }
  .st-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }
  .st-pill { display: flex; align-items: flex-start; gap: 10px; background: #f8fafc; border-radius: 8px; padding: 10px 12px; }
  .st-dot { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; margin-top: 3px; }
  .st-info { flex: 1; }
  .st-nome { font-size: 11px; color: #64748b; margin-bottom: 2px; }
  .st-val { font-size: 14px; font-weight: 500; color: #1e293b; }
  .st-cnt { font-size: 10px; color: #94a3b8; margin-top: 1px; }
  .mun-cols-head { display: grid; grid-template-columns: 1.8fr 1fr 1fr 1fr 1fr 24px; gap: 8px; padding: 6px 0 8px; font-size: 9px; font-weight: 600; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.04em; border-bottom: 1px solid #e2e8f0; margin-bottom: 4px; }
  .brow { border-bottom: 0.5px solid #f1f5f9; }
  .brow:last-child { border-bottom: none; }
  .bhead.mun-cols { display: grid; grid-template-columns: 1.8fr 1fr 1fr 1fr 1fr 24px; gap: 8px; align-items: center; padding: 10px 0; cursor: pointer; }
  .bhead:hover .bnome { color: #378ADD; }
  .bleft { display: flex; align-items: center; gap: 12px; flex: 1; min-width: 0; }
  .bav { width: 34px; height: 34px; border-radius: 50%; background: #E6F1FB; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: 500; color: #0C447C; flex-shrink: 0; }
  .binfo { flex: 1; min-width: 0; }
  .bnome { font-size: 13px; color: #1e293b; margin-bottom: 4px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; transition: color .15s; }
  .bmini-bar { height: 4px; background: #f1f5f9; border-radius: 2px; display: flex; overflow: hidden; }
  .bmini-seg { height: 100%; }
  .mun-col { text-align: right; }
  .mun-col-lbl { font-size: 8px; color: #cbd5e1; text-transform: uppercase; display: none; }
  .mun-col-val { font-size: 12px; font-weight: 600; color: #1e293b; }
  .bchevron { font-size: 18px; color: #cbd5e1; transition: transform .2s; flex-shrink: 0; line-height: 1; }
  details.brow[open] > summary .bchevron { transform: rotate(90deg); }
  .bdet { padding: 0 0 14px 46px; }
  details.st-grupo { margin-bottom: 6px; border-radius: 8px; overflow: hidden; border: 0.5px solid #e2e8f0; }
  details.st-grupo:last-child { margin-bottom: 0; }
  .st-grupo-hdr { display: flex; align-items: center; justify-content: space-between; padding: 8px 12px; cursor: pointer; gap: 10px; user-select: none; }
  .st-grupo-left { display: flex; align-items: center; gap: 8px; }
  .st-grupo-dot { width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0; }
  .st-grupo-nome { font-size: 12px; font-weight: 500; }
  .st-grupo-right { display: flex; align-items: center; gap: 10px; }
  .st-grupo-val { font-size: 12px; font-weight: 500; }
  .st-grupo-cnt { font-size: 10px; opacity: .7; }
  .st-grupo-chv { font-size: 16px; color: #94a3b8; transition: transform .2s; line-height: 1; }
  details.st-grupo[open] > summary .st-grupo-chv { transform: rotate(90deg); }
  .st-grupo-body { padding: 0 8px 8px; }
  .acao-lista { margin-top: 4px; }
  .acao-card { border-radius: 8px; padding: 10px 12px; margin-bottom: 6px; border: 0.5px solid transparent; }
  .acao-card:last-child { margin-bottom: 0; }
  .acao-top { display: flex; align-items: center; justify-content: space-between; margin-bottom: 4px; }
  .acao-tipo { font-size: 10px; font-weight: 600; text-transform: uppercase; letter-spacing: .04em; }
  .acao-valor { font-size: 11px; font-weight: 600; }
  .acao-titulo { font-size: 12px; color: #1e293b; line-height: 1.45; margin-bottom: 6px; }
  .acao-sep { height: 0.5px; margin-bottom: 7px; }
  .acao-row { display: flex; align-items: flex-start; flex-wrap: wrap; gap: 6px; }
  .acao-item { display: flex; flex-direction: column; gap: 2px; min-width: 0; }
  .acao-lbl { font-size: 10px; color: #94a3b8; white-space: nowrap; }
  .acao-v { font-size: 10px; color: #475569; font-weight: 500; max-width: 200px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .acao-div { font-size: 13px; color: #cbd5e1; padding: 0 2px; align-self: flex-end; padding-bottom: 1px; }
  .resumo-grid { display:grid; grid-template-columns:220px 190px 190px 220px; justify-content:center; column-gap:35px; }
  .resumo-item { background:transparent; padding:10px 10px; border-radius:10px; text-align: center; }
  .resumo-lbl { font-size: 11px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 6px; }
  .resumo-val { font-size: 20px; font-weight: 700; }
  .dp-head { display: grid; grid-template-columns: 1.6fr 1fr 1fr 1.3fr 1fr 1.3fr; gap: 8px; padding: 10px 8px; font-size: 10px; font-weight: 600; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.04em; border-bottom: 1px solid #e2e8f0; }
  details.dp-item-wrap { border-bottom: 0.5px solid #f1f5f9; }
  details.dp-item-wrap:last-child { border-bottom: none; }
  summary.dp-row { display: grid; grid-template-columns: 1.6fr 1fr 1fr 1.3fr 1fr 1.3fr; gap: 8px; padding: 12px 8px; align-items: center; cursor: pointer; list-style: none; }
  summary.dp-row::-webkit-details-marker { display: none; }
  summary.dp-row:hover { background: #fafbfc; }
  .dp-chevron { display: inline-block; font-size: 15px; color: #cbd5e1; margin-right: 6px; transition: transform .15s; }
  details.dp-item-wrap[open] > summary .dp-chevron { transform: rotate(90deg); }
  .dp-cell { font-size: 12px; }
  .dp-nome { font-weight: 600; color: #1e293b; }
  .dp-num { font-variant-numeric: tabular-nums; color: #475569; }
  .dp-detail { display: flex; flex-direction: column; gap: 4px; padding: 4px 12px 16px 30px; background: rgba(250,251,252,0.6); }
  details.dp-panel { background: #fff; border: 0.5px solid #e2e8f0; border-radius: 8px; }
  summary.dp-panel-hdr { display: flex; justify-content: space-between; align-items: center; cursor: pointer; padding: 10px 14px; font-size: 11px; font-weight: 600; color: #64748b; text-transform: uppercase; letter-spacing: 0.04em; list-style: none; }
  summary.dp-panel-hdr::-webkit-details-marker { display: none; }
  .dp-panel-chv { font-size: 15px; color: #cbd5e1; transition: transform .15s; }
  details.dp-panel[open] > summary .dp-panel-chv { transform: rotate(90deg); }
  details.dp-panel[open] > summary.dp-panel-hdr { border-bottom: 0.5px solid #f1f5f9; }
  .dp-panel-body { padding: 10px 14px 14px; }
  .dp-item { background: #fff; border: 0.5px solid #e2e8f0; border-radius: 6px; padding: 7px 10px; margin-bottom: 6px; }
  .dp-item:last-child { margin-bottom: 0; }
  .dp-item-top { display: flex; justify-content: space-between; gap: 8px; margin-bottom: 3px; }
  .dp-item-dest { font-size: 11px; color: #1e293b; line-height: 1.35; }
  .dp-item-val { font-size: 11px; font-weight: 600; color: #1e293b; white-space: nowrap; }
  .dp-item-meta { display: flex; align-items: center; gap: 6px; flex-wrap: wrap; }
  .dp-item-obs { font-size: 10px; color: #94a3b8; }
  .dp-vazio { font-size: 11px; color: #94a3b8; font-style: italic; padding: 8px 0; }
  .legenda-sit { display: flex; flex-wrap: wrap; gap: 10px; font-size: 10px; color: #64748b; margin-top: 10px; }
  .leg-dot { width: 8px; height: 8px; border-radius: 50%; display: inline-block; margin-right: 4px; vertical-align: middle; }
  .nota { font-size: 10px; color: #94a3b8; line-height: 1.5; }
  .app-header { position: relative; z-index: 2; max-width: 1180px; margin: 0 auto; padding: 22px 20px 4px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 12px; }
  .app-title { font-size: 20px; font-weight: 700; letter-spacing: 0.03em; color: #0C447C; text-transform: uppercase; }
  .nav-btn { display: inline-flex; align-items: center; gap: 6px; font-size: 12px; font-weight: 600; padding: 8px 16px; border-radius: 20px; border: 1px solid #378ADD; background: #fff; color: #0C447C; cursor: pointer; transition: background .15s, color .15s; text-decoration: none; }
  .nav-btn:hover { background: #378ADD; color: #fff; }
  .nav-btn.voltar { border-color: #cbd5e1; color: #5F5E5A; }
  .nav-btn.voltar:hover { background: #5F5E5A; border-color: #5F5E5A; color: #fff; }
  .view { position: relative; z-index: 1; display: none; }
  .view.ativa { display: block; }
  @media (max-width: 800px) {
    .kpi-grid { grid-template-columns: 1fr 1fr; }
    .two-col { grid-template-columns: 1fr; }
    .st-grid { grid-template-columns: 1fr; }
    .resumo-grid { grid-template-columns: 1fr 1fr; }
    .dp-head, summary.dp-row { grid-template-columns: 1.4fr 1fr 1fr; }
    .dp-head > div:nth-child(4), .dp-head > div:nth-child(5), .dp-head > div:nth-child(6),
    summary.dp-row > div:nth-child(4), summary.dp-row > div:nth-child(5), summary.dp-row > div:nth-child(6) { display: none; }
    .mun-cols-head, .bhead.mun-cols { grid-template-columns: 1.6fr 1fr 1fr 20px; }
    .mun-cols-head > div:nth-child(4), .bhead.mun-cols > .mun-col:nth-of-type(3) { display: none; }
    .app-header { justify-content: center; text-align: center; }
  }
"""


import base64
import os

def carregar_mapa_base64():
    """Lê o mapa de fundo (assets/sudeste_map.webp, ao lado deste script)
    e devolve como data URI, para embutir o HTML e não depender de a
    pasta assets/ estar junto do arquivo publicado."""
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'assets', 'sudeste_map.webp')
    with open(caminho, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode('ascii')
    return f"data:image/webp;base64,{b64}"


def montar_html(dashboard_body, depara_body, atualizado_em):
    mapa_uri = carregar_mapa_base64()
    css_final = CSS.replace("url('assets/sudeste_map.webp')", f"url('{mapa_uri}')")
    return f"""<!DOCTYPE html>
<html lang='pt-BR'>
<head>
<meta charset='UTF-8'>
<meta name='viewport' content='width=device-width, initial-scale=1.0'>
<title>Plano de Investimento — Macrorregião Sudeste</title>
<style>
{css_final}
</style>
</head>
<body>

<div class='app-header'>
  <div class='app-title'>Plano de Investimento — Macrorregião Sudeste</div>
  <div id='nav-dashboard'>
    <button class='nav-btn' onclick="mostrarPagina('depara')">Ver De/Para →</button>

  </div>
  <div id='nav-depara' style='display:none'>
    <button class='nav-btn voltar' onclick="mostrarPagina('dashboard')">← Voltar ao início</button>
  </div>
</div>

<div id='view-dashboard' class='view ativa'>
<div class='page'>
{dashboard_body}
</div>
</div>

<div id='view-depara' class='view'>
<div class='page'>
{depara_body}
</div>
</div>

<div style='text-align:center;font-size:10px;color:#94a3b8;padding:10px 0 24px;'>Última atualização: {atualizado_em} · dados da planilha de Monitoramento Fonte 95.</div>

<script>
function mostrarPagina(nome) {{
  document.getElementById('view-dashboard').classList.toggle('ativa', nome === 'dashboard');
  document.getElementById('view-depara').classList.toggle('ativa', nome === 'depara');
  document.getElementById('nav-dashboard').style.display = nome === 'dashboard' ? 'block' : 'none';
  document.getElementById('nav-depara').style.display = nome === 'depara' ? 'block' : 'none';
  window.scrollTo(0, 0);
}}
</script>

</body>
</html>"""


def main():
    if len(sys.argv) < 3:
        print("Uso: python3 build_dashboard.py CAMINHO.xlsx SAIDA.html")
        sys.exit(1)
    caminho_xlsx, saida = sys.argv[1], sys.argv[2]

    # GitHub Actions roda em UTC — ajusta para o horário de Brasília (UTC-3,
    # sem horário de verão desde 2019) para o rodapé exibir a hora certa.
    agora_br = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(hours=3)
    data_ref = agora_br.strftime('%d/%m/%Y')
    atualizado_em = agora_br.strftime('%d/%m/%Y às %H:%M')

    cons, plano_micro = carregar_dados(caminho_xlsx)
    dashboard_body = gerar_dashboard(cons, plano_micro, data_ref)
    depara_body = gerar_depara(cons, plano_micro, data_ref)
    html = montar_html(dashboard_body, depara_body, atualizado_em)

    with open(saida, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"OK — {saida} gerado com sucesso ({len(html)} bytes).")


if __name__ == '__main__':
    main()
